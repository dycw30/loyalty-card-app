from openpyxl import load_workbook

@app.route("/export")
def export():
    if not session.get("logged_in"):
        return redirect(url_for("login"))

    # Load orders from database
    conn = sqlite3.connect(DB_NAME)
    df = pd.read_sql_query("SELECT * FROM orders", conn)
    conn.close()

    # Prepare additional columns matching your structure
    df["Today's Order"] = 1
    df["Redeemed"] = df["redeemed"].apply(lambda x: 1 if x else 0)

    # Load your existing .xlsm file with formulas intact
    template_file = "Bound Cafe Test 3.xlsm"
    wb = load_workbook(template_file, keep_vba=True)
    ws = wb["Sheet1"]

    # Insert starting at A2 downward
    start_row = 2
    for i, row in df.iterrows():
        ws.cell(row=start_row + i, column=1, value=row["customer_name"])    # A
        # B left blank for manual Unique ID
        ws.cell(row=start_row + i, column=3, value=1)                       # C - Today's Order
        ws.cell(row=start_row + i, column=4, value=row["drink_type"])       # D
        # E left for your Excel formula (total orders)
        ws.cell(row=start_row + i, column=6, value=row["tokens"])           # F - Tokens
        ws.cell(row=start_row + i, column=7, value=f"{row['Redeemed']} / {row['tokens']}") # G

    # Save to a new file so you keep original safe
    export_file = "Bound Cafe Exported.xlsm"
    wb.save(export_file)
    return send_file(export_file, as_attachment=True)
